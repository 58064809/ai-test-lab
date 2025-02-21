# coding=utf-8
# 2024/11/26 17:25
from datetime import datetime

import requests
from pypinyin import lazy_pinyin

from other_settings import Connection
from openpyxl import load_workbook
from util.log_handle import log
from openpyxl.styles import PatternFill
import traceback


class ShenJi:
    def __init__(self):
        self.env = 'shenji'
        self.connect, self.cursor = Connection.get_db_connect(self.env)
        self.filename = r"C:\Users\ll07\Desktop\汇总V2_20250116.xlsx"
        self.workbook = load_workbook(self.filename)
        self.host = 'https://oasd.lianlianlvyou.com'
        self.session = requests.session()
        self.phone = '13550087714'
        self.password = '087714'

    def get_data(self):
        all_sheets_data = {}
        # 遍历每一个工作表
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]

            # 取第2行作为表头
            headers = [cell.value for cell in ws[2]]

            # 构建数据列表
            sheet_data = []
            for row in ws.iter_rows(min_row=3, values_only=True):  # 从第3行开始
                if any(cell for cell in row):
                    row_dict = dict(zip(headers, row))
                    sheet_data.append(row_dict)

            all_sheets_data[sheet_name] = sheet_data
        return all_sheets_data

    def insert_store(self):
        # 打印每个工作表的所有数据
        all_sheets_data = {sheet: data for sheet, data in self.get_data().items() if sheet == "需求"}
        num = 1
        for sheet, data in all_sheets_data.items():
            log.success(f"工作表 {sheet} 的数据:")
            for item in data:
                log.info(item)
                if isinstance(item['openedDate'], str):
                    item['openedDate'] = item['openedDate'].replace('/', '-')
                    item['openedDate'] = datetime.strptime(item['openedDate'], '%Y-%m-%d %H:%M:%S')
                if isinstance(item['reviewedDate'], str):
                    item['reviewedDate'] = item['reviewedDate'].replace('/', '-')
                    item['reviewedDate'] = datetime.strptime(item['reviewedDate'], '%Y-%m-%d %H:%M:%S')
                if isinstance(item['closedDate'], str):
                    item['closedDate'] = item['closedDate'].replace('/', '-')
                    item['closedDate'] = datetime.strptime(item['closedDate'], '%Y-%m-%d %H:%M:%S')
                if int(item['product']) == 8:
                    item['project'] = 12
                    reviewedBy = '叶忠海'
                elif int(item['product']) == 7:
                    item['project'] = 10
                    reviewedBy = '高恺'
                else:
                    item['project'] = 4
                    reviewedBy = '张跃建'
                if item['status'] == 'active':
                    item['closedBy'] = ""
                    item['closedDate'] = None

                sql = "INSERT INTO `zentao`.`zt_story`(product,title,pri,status,stage,openedBy,openedDate,assignedTo,assignedDate,lastEditedBy,lastEditedDate,reviewedBy,reviewedDate,closedBy,closedDate,closedReason) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);"
                self.cursor.execute(sql, (
                    item['product'], item['title'], item['pri'], item['status'],
                    'developing' if item['status'] == 'active' else item['status'], item['openedBy'],
                    item['openedDate'], item['status'], item['closedDate'], item['closedBy'],
                    item['closedDate'], item['reviewedBy'], item['reviewedDate'], item['closedBy'], item['closedDate'],
                    "done"))
                last_id = self.cursor.lastrowid

                sql = "INSERT INTO `zentao`.`zt_projectstory`(`project`,`product`,`branch`,`story`,`version`,`order`) VALUES (%s,%s,%s,%s,%s,%s);"
                self.cursor.execute(sql, (item['project'], item['product'], 0, last_id, 1, num))

                sql = "INSERT INTO `zentao`.`zt_action`(`objectType`, `objectID`, `product`, `project`, `execution`, `actor`, `action`, `date`, `comment`, `extra`, `read`, `vision`, `efforted`) VALUES (%s, %s,%s,%s, %s,%s, %s, %s, %s, %s, %s, %s, %s);"
                self.cursor.execute(sql, (
                    'story', last_id, ',' + str(item['product']) + ',', item['project'], 0, item['openedBy'],
                    'assigned',
                    item['reviewedDate'], '', ''.join(lazy_pinyin(reviewedBy)), '0', 'rnd', 0))

                sql = "INSERT INTO `zentao`.`zt_storyspec`(`story`,`version`,`title`,`spec`) VALUES (%s,%s,%s,%s);"
                self.cursor.execute(sql, (last_id, 1, item['title'], item['spec']))

                if item['status'] != 'active':
                    sql = "INSERT INTO `zentao`.`zt_action`(`objectType`, `objectID`, `product`, `project`, `execution`, `actor`, `action`, `date`, `comment`, `extra`, `read`, `vision`, `efforted`) VALUES (%s, %s,%s,%s, %s,%s, %s, %s, %s, %s, %s, %s, %s);"
                    self.cursor.execute(sql, (
                        'story', last_id, ',' + str(item['product']) + ',', item['project'], 0, item['closedBy'],
                        'closed',
                        item['closedDate'], item['comment'], "Done|draft", '0', 'rnd', 0))
                    action_id = self.cursor.lastrowid

                    sql = "INSERT INTO `zentao`.`zt_history`(`action`, `field`, `old`, `new`, `diff`) VALUES (%s, %s, %s, %s, %s);"
                    self.cursor.execute(sql, [action_id, "closedBy", "", item['closedBy'], ""])
                num += 1

    def insert_bug(self):
        all_sheets_data = {sheet: data for sheet, data in self.get_data().items() if sheet.upper() == "BUG"}
        for sheet, data in all_sheets_data.items():
            log.success(f"工作表 {sheet} 的数据:")
            for item in data:
                log.info(item)
                if isinstance(item['openedDate'], str):
                    item['openedDate'] = item['openedDate'].replace('/', '-')
                    item['openedDate'] = datetime.strptime(item['openedDate'], '%Y-%m-%d %H:%M:%S')
                if isinstance(item['resolvedDate'], str):
                    item['resolvedDate'] = item['resolvedDate'].replace('/', '-')
                    item['resolvedDate'] = datetime.strptime(item['resolvedDate'], '%Y-%m-%d %H:%M:%S')
                if isinstance(item['closedDate'], str):
                    item['closedDate'] = item['closedDate'].replace('/', '-')
                    item['closedDate'] = datetime.strptime(item['closedDate'], '%Y-%m-%d %H:%M:%S')
                if item['status'] == 'active':
                    item['resolvedBy'] = ""
                    item['resolvedDate'] = None
                    item['closedBy'] = ""
                    item['closedDate'] = None
                sql = "INSERT INTO `zentao`.`zt_bug`(`project`, `product`, `storyVersion`,`title`, `severity`, `pri`, `type`, `steps`, `status`,`confirmed`, `openedBy`, `openedDate`, `openedBuild`, `assignedTo`, `assignedDate`, `deadline`, `resolvedBy`, `resolution`, `resolvedBuild`, `resolvedDate`, `closedBy`, `closedDate`,`linkBug`,`lastEditedBy`, `lastEditedDate`) VALUES (%s,%s,%s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s,%s, %s,%s, %s, %s, %s,%s,%s, %s);"
                self.cursor.execute(sql, (
                    item['project'], item['product'], 1, item['title'], item['severity'], item['pri'], 'codeerror',
                    item['steps'],
                    item['status'], 1, item['openedBy'], item['openedDate'], 'trunk', 'closed', item['closedDate'],
                    '0000-00-00', item['resolvedBy'],
                    'fixed', 'trunk', item['resolvedDate'], item['closedBy'], item['closedDate'], 0,
                    item['closedBy'], item['closedDate']))

                last_id = self.cursor.lastrowid
                sql = "INSERT INTO `zentao`.`zt_action`(`objectType`, `objectID`, `product`, `project`, `execution`, `actor`, `action`, `date`, `comment`, `extra`, `read`, `vision`, `efforted`) VALUES (%s, %s,%s,%s, %s,%s, %s, %s, %s, %s, %s, %s, %s);"
                self.cursor.execute(sql, (
                    'bug', last_id, ',' + str(item['product']) + ',', item['project'], 0, item['openedBy'], 'opened',
                    item['openedDate'], '', '', '1', 'rnd', 0))

                sql = "INSERT INTO `zentao`.`zt_action`(`objectType`, `objectID`, `product`, `project`, `execution`, `actor`, `action`, `date`, `comment`, `extra`, `read`, `vision`, `efforted`) VALUES (%s, %s,%s,%s, %s,%s, %s, %s, %s, %s, %s, %s, %s);"
                self.cursor.execute(sql, (
                    'bug', last_id, ',' + str(item['product']) + ',', item['project'], 0, item['resolvedBy'],
                    'resolved',
                    item['resolvedDate'], '', 'fixed', '1', 'rnd', 0))

                sql = "INSERT INTO `zentao`.`zt_action`(`objectType`, `objectID`, `product`, `project`, `execution`, `actor`, `action`, `date`, `comment`, `extra`, `read`, `vision`, `efforted`) VALUES (%s, %s,%s,%s, %s,%s, %s, %s, %s, %s, %s, %s, %s);"
                self.cursor.execute(sql, (
                    'bug', last_id, ',' + str(item['product']) + ',', item['project'], 0, item['closedBy'], 'closed',
                    item['closedDate'], '已解决', '', '0', 'rnd', 0))
                action_id = self.cursor.lastrowid

                sql = "INSERT INTO `zentao`.`zt_history`(`action`, `field`, `old`, `new`, `diff`) VALUES (%s, %s, %s, %s, %s);"
                self.cursor.execute(sql, [action_id, "closedBy", "", item['closedBy'], ""])

    def login(self):
        url = self.host + '/common/open/login/loginByPassword'
        data = {"phoneNumber": self.phone, "password": self.password}
        resp = self.session.post(url, json=data).json()
        if resp['code'] == 200:
            log.success(f'登录成功，token:{resp["data"]["token"]}')
            return resp['data']['token']
        raise Exception('登录失败')

    def apply(self, token):
        all_sheets_data = {sheet: data for sheet, data in self.get_data().items() if
                           sheet != "需求" and sheet.upper() != "BUG"}
        aggregated_data = []
        for sheet, records in all_sheets_data.items():
            for i, record in enumerate(records):
                # 处理字符串时间为 datetime 格式（如果需要）
                if isinstance(record['申请时间'], str):
                    record['申请时间'] = record['申请时间'].replace('/', '-')
                    record['申请时间'] = datetime.strptime(record['申请时间'], '%Y-%m-%d %H:%M:%S')
                # 记录所属的申请类型
                record['sheet'] = sheet
                record['i'] = i
                aggregated_data.append(record)
        sorted_data = sorted(aggregated_data, key=lambda x: x['申请时间'])
        for item in sorted_data:
            log.success(f"工作表 {item['sheet']} 的数据:")
            sheet_obj = self.workbook[item['sheet']]
            max_column = max((i for i, cell in enumerate(sheet_obj[2], start=1) if cell.value not in (None, '')),
                             default=0)
            log.info(item)
            if item['sheet'] == "发布申请":
                data = {"list": [
                    {"content": "\"%s\"" % item['申请人'], "field": "mates0", "fieldName": "申请人",
                     "tag": "mates"}, {
                        "content": "{\"id\":16751,\"parentId\":null,\"deptCode\":\"D00000600\",\"deptName\":\"微信事业部\",\"fullDeptName\":\"微信事业部\",\"approvalCompanyId\":null,\"companyName\":null,\"children\":null,\"positionName\":null,\"positionId\":null}",
                        "field": "chooseDepartment1", "fieldName": "申请人部门", "tag": "chooseDepartment"},
                    {"content": "\"\"", "field": "tooltips4", "fieldName": "说明文字", "tag": "tooltips"},
                    {"content": "\"%s\"" % item['需求目标'], "field": "textarea2", "fieldName": "需求目标",
                     "tag": "textarea"},
                    {"content": "\"\"", "field": "tooltips6", "fieldName": "说明文字", "tag": "tooltips"},
                    {"content": "\"%s\"" % item['上线功能点'], "field": "textarea3",
                     "fieldName": "上线功能点", "tag": "textarea"},
                    {"content": "\"\"", "field": "tooltips8", "fieldName": "说明文字", "tag": "tooltips"},
                    {"content": "\"%s\"" % item['上线内容'], "field": "textarea7",
                     "fieldName": "上线内容", "tag": "textarea"},
                    {"content": "\"\"", "field": "textarea9", "fieldName": "执行SQL", "tag": "textarea"}, {
                        "content": "\"%s\"" % item['测试报告文件'],
                        "field": "file5", "fieldName": "测试报告附件", "tag": "file"}], "userId": 42082,
                    "userName": "路昭凡(895110)", "companyId": 1, "templateId": "470"}
                range_num = 4
            elif item['sheet'] == "变更申请":
                data = {"list": [
                    {"content": "\"%s\"" % item['申请人'], "field": "mates0", "fieldName": "申请人",
                     "tag": "mates"},
                    {
                        "content": "{\"id\":16751,\"parentId\":null,\"deptCode\":\"D00000600\",\"deptName\":\"微信事业部\",\"fullDeptName\":\"微信事业部\",\"approvalCompanyId\":null,\"companyName\":null,\"children\":null,\"positionName\":null,\"positionId\":null}",
                        "field": "chooseDepartment1", "fieldName": "所属部门", "tag": "chooseDepartment"},
                    {"content": "\"%s\"" % item['申请/变更理由'], "field": "textarea3",
                     "fieldName": "申请/变更理由",
                     "tag": "textarea"},
                    {"content": "\"%s\"" % item['申请/变更权限内容'], "field": "textarea2",
                     "fieldName": "申请/变更权限内容",
                     "tag": "textarea"}], "userId": 42082, "userName": "路昭凡(895110)", "companyId": 1,
                    "templateId": "471"}
                range_num = 3
            elif item['sheet'] == "权限审阅":
                data = {"list": [
                    {"content": "\"%s\"" % item['申请人'], "field": "mates0", "fieldName": "发起人",
                     "tag": "mates"}, {
                        "content": "{\"id\":16751,\"parentId\":null,\"deptCode\":\"D00000600\",\"deptName\":\"微信事业部\",\"fullDeptName\":\"微信事业部\",\"approvalCompanyId\":null,\"companyName\":null,\"children\":null,\"positionName\":null,\"positionId\":null}",
                        "field": "chooseDepartment1", "fieldName": "所属部门", "tag": "chooseDepartment"},
                    {"content": "\"%s\"" % item['审阅内容'], "field": "textarea2", "fieldName": "审阅内容",
                     "tag": "textarea"}, {
                        "content": "\"%s\"" % item['附件上传'] if item['附件上传'] else "\"\"",
                        "field": "file3", "fieldName": "附件上传", "tag": "file"},
                    {"content": "\"%s\"" % item['备注'] if item['备注'] else "\"\"", "field": "textarea4",
                     "fieldName": "问题备注",
                     "tag": "textarea"}],
                    "userId": 42082, "userName": "路昭凡(895110)", "companyId": 1, "templateId": "474"}
                range_num = 1
            elif item['sheet'] == "离职权限关闭申请":
                data = {"list": [
                    {"content": "\"%s\"" % item['申请人'], "field": "mates0", "fieldName": "发起人",
                     "tag": "mates"}, {
                        "content": "{\"id\":16751,\"parentId\":null,\"deptCode\":\"D00000600\",\"deptName\":\"微信事业部\",\"fullDeptName\":\"微信事业部\",\"approvalCompanyId\":null,\"companyName\":null,\"children\":null,\"positionName\":null,\"positionId\":null}",
                        "field": "chooseDepartment1", "fieldName": "所属部门", "tag": "chooseDepartment"},
                    {"content": "[\"%s\"]" % item['关闭类型'], "field": "radio-group5", "fieldName": "关闭类型",
                     "tag": "radio-group"},
                    {"content": "\"\"", "field": "tooltips4", "fieldName": "说明文字", "tag": "tooltips"},
                    {"content": "\"%s\"" % item['关闭范围'], "field": "textarea2", "fieldName": "账号关闭范围",
                     "tag": "textarea"}], "userId": 42082, "userName": "路昭凡(895110)", "companyId": 1,
                    "templateId": "475"}
                range_num = 3
            elif item['sheet'] == "金蝶账号关闭":
                data = {"list": [
                    {"content": "\"%s\"" % item['申请人'], "field": "mates0", "fieldName": "发起人", "tag": "mates"}, {
                        "content": "{\"id\":12,\"parentId\":null,\"deptCode\":\"E000000000000\",\"deptName\":\"财务中心\",\"fullDeptName\":\"财务中心\",\"approvalCompanyId\":null,\"companyName\":null,\"children\":null,\"positionName\":null,\"positionId\":null}",
                        "field": "chooseDepartment1", "fieldName": "所属部门", "tag": "chooseDepartment"},
                    {"content": "\"\"", "field": "tooltips4", "fieldName": "说明文字", "tag": "tooltips"},
                    {"content": "\"%s\"" % item['关闭原因'], "field": "textarea2", "fieldName": "关闭原因",
                     "tag": "textarea"},
                    {"content": "\"%s\"" % item['关闭账号'], "field": "textarea6", "fieldName": "关闭账号",
                     "tag": "textarea"}],
                    "userId": 42082, "userName": "路昭凡(895110)", "companyId": 1, "templateId": "483"}
                range_num = 3
            elif item['sheet'] == "金蝶账号开通":
                data = {"list": [
                    {"content": "\"%s\""%item['申请人'], "field": "mates0", "fieldName": "申请人", "tag": "mates"}, {
                        "content": "{\"id\":12,\"parentId\":null,\"deptCode\":\"E000000000000\",\"deptName\":\"财务中心\",\"fullDeptName\":\"财务中心\",\"approvalCompanyId\":null,\"companyName\":null,\"children\":null,\"positionName\":null,\"positionId\":null}",
                        "field": "chooseDepartment1", "fieldName": "所属部门", "tag": "chooseDepartment"},
                    {"content": "\"%s\""%item['申请理由'], "field": "textarea2", "fieldName": "申请理由", "tag": "textarea"},
                    {"content": "\"%s\""%item['账号信息'], "field": "textarea3", "fieldName": "账号信息", "tag": "textarea"}],
                        "userId": 42082, "userName": "路昭凡(895110)", "companyId": 1, "templateId": "482"}
                range_num = 3
            elif item['sheet'] == "业务需求申请":
                if item['申请人部门'] == "财务部门":
                    cid = 12
                    deptCode = "E000000000000"
                    deptName = "财务中心"
                    fullDeptName = "财务中心"
                elif item['申请人部门'] == "联联旅游":
                    cid = 15852
                    deptCode = "Ж000000000000"
                    deptName = "联联旅游"
                    fullDeptName = "联联旅游"
                elif item['申请人部门'] == "平台运营中心":
                    cid = 17178
                    deptCode = "D00001027"
                    deptName = "平台运营中心"
                    fullDeptName = "平台运营中心"
                elif item['申请人部门'] == "公域中台":
                    cid = 16532
                    deptCode = "D00000601"
                    deptName = "公域中台"
                    fullDeptName = "公域中台"
                else:
                    cid = 16751
                    deptCode = "D00000600"
                    deptName = "微信事业部"
                    fullDeptName = "微信事业部"
                data = {"list": [
                    {"content": "\"%s\"" % item['申请人'], "field": "mates0", "fieldName": "申请人",
                     "tag": "mates"}, {
                        "content": "{\"id\":%s,\"parentId\":null,\"deptCode\":\"%s\",\"deptName\":\"%s\",\"fullDeptName\":\"%s\",\"approvalCompanyId\":null,\"companyName\":null,\"children\":null,\"positionName\":null,\"positionId\":null}" % (
                            cid, deptCode, deptName, fullDeptName),
                        "field": "chooseDepartment1", "fieldName": "申请人部门", "tag": "chooseDepartment"},
                    {"content": "\"\"", "field": "tooltips4", "fieldName": "说明文字", "tag": "tooltips"},
                    {"content": "\"%s\"" % item['需求背景'], "field": "textarea2", "fieldName": "需求背景",
                     "tag": "textarea"},
                    {"content": "\"%s\"" % item['需求描述'], "field": "textarea3", "fieldName": "需求描述",
                     "tag": "textarea"},
                    {"content": "\"\"", "field": "tooltips6", "fieldName": "说明文字", "tag": "tooltips"}, {
                        "content": "\"%s\"" % item['需求文件'] if item['需求文件'] else "\"\"",
                        "field": "file5", "fieldName": "需求文件", "tag": "file"}], "userId": 42082,
                    "userName": "路昭凡(895110)", "companyId": 1, "templateId": "469"}
                range_num = 3
            else:
                data = {"list": [
                    {"content": "\"%s\"" % item['申请人'], "field": "mates0", "fieldName": "发起人", "tag": "mates"}, {
                        "content": "{\"id\":16751,\"parentId\":null,\"deptCode\":\"D00000600\",\"deptName\":\"微信事业部\",\"fullDeptName\":\"微信事业部\",\"approvalCompanyId\":null,\"companyName\":null,\"children\":null,\"positionName\":null,\"positionId\":null}",
                        "field": "chooseDepartment1", "fieldName": "发起部门", "tag": "chooseDepartment"},
                    {"content": "\"%s\"" % item['关闭类型'], "field": "radio2", "fieldName": "变更类型",
                     "tag": "radio"},
                    {"content": "\"%s\"" % item['执行SQL'], "field": "textarea3", "fieldName": "执行SQL",
                     "tag": "textarea"}],
                    "userId": 42082, "userName": "路昭凡(895110)", "companyId": 1, "templateId": "472"}
                range_num = 3
            url = self.host + '/it_administration/approval/apply'
            headers = {"authorization": token}
            resp = self.session.post(url, json=data, headers=headers).json()
            if resp['code'] == 200:
                SP = resp['data']
                log.success("创建审批流成功,单号:%s" % SP)
                url = self.host + '/it_administration/approval/admin/list?search=%s&isHandled=0&pageIndex=1&pageSize=10' % SP
                respo = self.session.get(url, headers=headers).json()
                if respo['data']['data'] != []:
                    sp_id = respo['data']['data'][0]['id']
                    log.success("获取审批流单号ID成功:%s" % sp_id)
                    for _ in range(range_num):
                        url = self.host + '/it_administration/approval/apply/admin/audit'
                        data = {"auditStatus": 1, "applyId": sp_id, "remark": "审批通过"}
                        response = self.session.post(url, json=data, headers=headers).json()
                        if response['code'] == 200:
                            log.success("审批通过")
                            default_fill = PatternFill()
                            sheet_obj.cell(row=int(item['i']) + 3, column=max_column+1, value=SP).fill = default_fill
                            self.workbook.save(self.filename)
                            log.success(f"在{item['sheet']}表第{max_column+1}列写入审批单号{SP}成功")
                        else:
                            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            sheet_obj.cell(row=int(item['i']) + 3, column=max_column+1, value="失败").fill = red_fill
                            self.workbook.save(self.filename)
                            log.error("审批失败")
                else:
                    log.error("获取审批流单号ID失败,审批单号为:%s" % SP)
            else:
                log.error("创建审批流失败")


if __name__ == '__main__':
    shenji = ShenJi()
    # try:
    #     shenji.connect.begin()
    #     # # 需求向
    #     shenji.insert_store()
    #     # BUG向
    #     shenji.insert_bug()
    #
    #     shenji.connect.commit()
    # except Exception as e:
    #     shenji.connect.rollback()
    #     traceback.print_exc()


    token = shenji.login()
    shenji.apply(token)
