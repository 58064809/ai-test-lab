# coding=utf-8
# 2024/8/1 下午3:09

from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\ll07\Desktop\com_user.xlsx')
sheet = wb['导出结果']
list1 = list(sheet.values)[1:]
wb.close()
wb = load_workbook(r'C:\Users\ll07\Desktop\花名册.xlsx')
sheet = wb['23年12月花名册']
li = list(sheet.values)[1:]
for j in range(len(li)):
    for i in range(len(list1)):
        if str(li[j][0]) == str(list1[i][1]):
            sheet.cell(row=j + 2, column=12).value = str(list1[i][-2])
            if str(li[j][10]) != str(list1[i][-2]):
                sheet.cell(row=j+2, column=13).value = "异常"
wb.save(r'C:\Users\ll07\Desktop\花名册.xlsx')


